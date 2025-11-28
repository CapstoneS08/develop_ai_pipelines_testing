#!/usr/bin/env python3
"""Summarize comments in an Excel file into positives and areas for improvement.

Reads comments from a specified input column (default B) and writes two summaries:
- Column C: Summary of positive aspects/strengths
- Column D: Summary of areas for improvement/weaknesses

If the environment variable OPENAI_API_KEY is set, the script will call the OpenAI
Chat API to produce intelligent summaries; otherwise it uses a small local heuristic.

Usage examples in README.md. Safe defaults: read active sheet, start at row 1,
save result to a new file with suffix _summarized.xlsx.
"""
import argparse
import os
import time
from typing import Optional, Tuple

try:
    from dotenv import load_dotenv
    load_dotenv()  # Load .env file if it exists
except ImportError:
    pass  # dotenv not installed, skip

try:
    import openai
except Exception:
    openai = None

from openpyxl import load_workbook


def summarize_local(text: str) -> Tuple[str, str]:
    """Very small fallback summarizer: extracts positives and areas for improvement.
    
    Returns (positives_summary, improvements_summary)
    """
    if not text:
        return "", ""
    
    s = str(text).strip().lower()
    
    # Enhanced keyword-based heuristic with stronger positive signals
    strong_positive_indicators = ["great", "excellent", "love", "amazing", "wonderful", 
                                  "delighted", "fantastic", "outstanding", "passionate",
                                  "pleased", "satisfied", "happy", "thank you", "thanks"]
    positive_indicators = ["good", "helpful", "strength", "positive", "well", "best", 
                          "appreciate", "clear", "transparent", "flexible", "adaptable",
                          "smooth", "reliable", "quality", "professional"]
    negative_indicators = ["bad", "poor", "improve", "should", "could be better",
                          "issue", "problem", "concern", "weak", "lacking", "difficult",
                          "late", "delay", "not", "disappointing", "unfortunately"]
    
    # If the overall text is overwhelmingly positive, treat it as such
    strong_positive_count = sum(1 for word in strong_positive_indicators if word in s)
    negative_count = sum(1 for word in negative_indicators if word in s)
    
    # If highly positive with no negatives, put everything in positives
    if strong_positive_count >= 2 and negative_count == 0:
        summary = text.strip()
        if len(summary) > 100:
            summary = summary[:97] + "..."
        return summary, "None mentioned"
    
    sentences = []
    for sep in [". ", "! ", "? ", "\n"]:
        if sep in text:
            sentences = [s.strip() for s in text.split(sep) if s.strip()]
            break
    
    if not sentences:
        sentences = [text]
    
    positives = []
    improvements = []
    
    for sentence in sentences:
        sent_lower = sentence.lower()
        
        # Check for strong positive signals first
        has_strong_positive = any(word in sent_lower for word in strong_positive_indicators)
        has_positive = any(word in sent_lower for word in positive_indicators)
        has_negative = any(word in sent_lower for word in negative_indicators)
        
        # Prioritize strong positives
        if has_strong_positive and not has_negative:
            positives.append(sentence)
        elif has_positive and not has_negative:
            positives.append(sentence)
        elif has_negative and not has_strong_positive:
            improvements.append(sentence)
        elif not positives and not improvements:
            # Default to positives if unclear and no clear negatives
            positives.append(sentence)
    
    pos_summary = ". ".join(positives[:3]) if positives else "None mentioned"
    imp_summary = ". ".join(improvements[:2]) if improvements else "None mentioned"
    
    # Truncate if too long
    if len(pos_summary) > 150:
        pos_summary = pos_summary[:147] + "..."
    if len(imp_summary) > 150:
        imp_summary = imp_summary[:147] + "..."
    
    return pos_summary, imp_summary


def summarize_with_openai(text: str, model: str = "gpt-4o") -> Optional[Tuple[str, str]]:
    """Call OpenAI ChatCompletion to extract positives and areas for improvement.
    
    Returns (positives_summary, improvements_summary) or None if failed.
    Requires openai package and OPENAI_API_KEY env var set.
    """
    if openai is None:
        return None
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return None
    
    prompt = """Analyze the following comment carefully and categorize the feedback into four different categories and two distinct areas:

The four categories are:
1. Product
2. Service
3. Delivery
4. Payment

The two areas to focus on are:
1. POSITIVES: Extract ONLY genuinely positive feedback - things the person praised, appreciated, or said were done well. If the comment is purely critical or neutral, write "None mentioned".

2. IMPROVEMENTS: Extract ONLY constructive criticism, issues raised, or areas needing improvement. If the comment is purely positive with no concerns, write "None mentioned".

CRITICAL RULES:
- Do NOT convert positive statements into improvement areas
- Do NOT assume there must be problems if only praise is given
- Do NOT assume there must be positives if only criticism is given
- If a comment is entirely positive (e.g., "Great service, very satisfied"), put it in POSITIVES and put "None mentioned" for IMPROVEMENTS
- If a comment is entirely about problems (e.g., "Delivery was late, poor quality"), put "None mentioned" for POSITIVES

Format your response EXACTLY as:
POSITIVES: [your summary or "None mentioned"]
IMPROVEMENTS: [your summary or "None mentioned"]

Comment:
""" + str(text)
    
    try:
        # Use modern OpenAI API (v1.0+)
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=200,
            temperature=0.2,
        )
        content = resp.choices[0].message.content.strip()
        
        # Parse response
        positives = ""
        improvements = ""
        
        for line in content.split("\n"):
            line = line.strip()
            if line.startswith("POSITIVES:"):
                positives = line.replace("POSITIVES:", "").strip()
            elif line.startswith("IMPROVEMENTS:"):
                improvements = line.replace("IMPROVEMENTS:", "").strip()
        
        return positives, improvements
    except Exception as e:
        print(f"OpenAI call failed: {e}")
        return None


def run(
    path: str,
    sheet_name: Optional[str],
    input_col: str,
    positives_col: str,
    improvements_col: str,
    start_row: int,
    model: str,
    overwrite: bool,
    delay: float,
    force: bool = False,
):
    # Show which mode we're using
    api_key = os.getenv("OPENAI_API_KEY")
    if api_key and openai is not None:
        print(f"🤖 Using OpenAI API (model: {model})")
        print(f"   API Key: {api_key[:10]}...{api_key[-4:]}")
    else:
        print("⚠️  Using LOCAL summarizer (OpenAI not available)")
        if not api_key:
            print("   Reason: OPENAI_API_KEY not found in environment")
        if openai is None:
            print("   Reason: openai package not installed")
    print()
    
    wb = load_workbook(path)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    max_row = ws.max_row
    changed = 0
    for r in range(start_row, max_row + 1):
        in_cell = f"{input_col}{r}"
        pos_cell = f"{positives_col}{r}"
        imp_cell = f"{improvements_col}{r}"
        
        comment = ws[in_cell].value
        if comment is None or str(comment).strip() == "":
            continue
        
        # Skip if already has both output cells populated (preserve existing)
        if not force and ws[pos_cell].value not in (None, "") and ws[imp_cell].value not in (None, ""):
            continue

        positives = ""
        improvements = ""
        
        # Prefer OpenAI when available
        if os.getenv("OPENAI_API_KEY") and openai is not None:
            result = summarize_with_openai(comment, model=model)
            if result:
                positives, improvements = result
            else:
                # Fallback to local if OpenAI fails
                positives, improvements = summarize_local(comment)
        else:
            positives, improvements = summarize_local(comment)

        ws[pos_cell].value = positives
        ws[imp_cell].value = improvements
        changed += 1
        print(f"Row {r}: Positives ({len(positives)} chars) | Improvements ({len(improvements)} chars)")
        
        if delay and os.getenv("OPENAI_API_KEY"):
            time.sleep(delay)

    if changed == 0:
        print("No comments found or no changes needed.")

    # Write output file
    dirname, basename = os.path.split(path)
    name, ext = os.path.splitext(basename)
    if overwrite:
        out_path = path
    else:
        out_path = os.path.join(dirname, f"{name}_summarized{ext}")
    
    # Try to save with helpful error message if file is locked
    try:
        wb.save(out_path)
        print(f"Saved {out_path} ({changed} rows updated)")
    except PermissionError:
        print(f"\n❌ ERROR: Cannot save to '{out_path}'")
        print("   The file is currently open in another program (Excel, etc.)")
        print("\n   Solutions:")
        print("   1. Close the Excel file and run the script again")
        print("   2. Run without --overwrite to create a new file instead:")
        print(f"      python scripts\\summarize_excel.py \"{path}\" --start-row {start_row}")
        raise


def parse_args():
    p = argparse.ArgumentParser(description="Summarize comments into positives and areas for improvement")
    p.add_argument("file", help="Path to the Excel file (.xlsx)")
    p.add_argument("--sheet", help="Sheet name (default: active)")
    p.add_argument("--input-col", default="B", help="Column with comments (default B)")
    p.add_argument("--positives-col", default="C", help="Column for positives summary (default C)")
    p.add_argument("--improvements-col", default="D", help="Column for improvements summary (default D)")
    p.add_argument("--start-row", type=int, default=1, help="Row to start reading from (default 1)")
    p.add_argument("--model", default="gpt-4o", help="OpenAI model to use when available")
    p.add_argument("--overwrite", action="store_true", help="Overwrite the input file instead of creating _summarized")
    p.add_argument("--force", action="store_true", help="Force reprocessing of all rows, even if already summarized")
    p.add_argument("--delay", type=float, default=1.0, help="Seconds delay between OpenAI calls (default 1.0)")
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    try:
        run(
            path=args.file,
            sheet_name=args.sheet,
            input_col=args.input_col.upper(),
            positives_col=args.positives_col.upper(),
            improvements_col=args.improvements_col.upper(),
            start_row=args.start_row,
            model=args.model,
            overwrite=args.overwrite,
            delay=args.delay,
            force=args.force,
        )
    except Exception as e:
        print(f"Error: {e}")
        raise
