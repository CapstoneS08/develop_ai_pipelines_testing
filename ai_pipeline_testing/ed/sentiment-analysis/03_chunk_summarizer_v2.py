#!/usr/bin/env python3
"""Summarize comments in an Excel file into positives and areas for improvement with sentiment analysis.

Flow:
1. LLM splits comment into positive and negative parts
2. Each part is chunked/stringified separately
3. Sentiment analysis is performed on each chunk
4. Overall sentiment is analyzed on the original full comment
5. Outputs:
   - Column C: Summary of positive aspects
   - Column D: Summary of areas for improvement
   - Column E: Sentiment score for positives (1-5)
   - Column F: Sentiment score for improvements (1-5)
   - Column G: Overall sentiment score based on categorized chunks (1-5)
   - Column H: Raw sentiment score from original comment only (1-5)
   - Column I: Comment category/categories

If the environment variable OPENAI_API_KEY is set, the script will call the OpenAI
Chat API to produce intelligent summaries; otherwise it uses a small local heuristic.

Usage examples in README.md. Safe defaults: read active sheet, start at row 1,
save result to a new file with suffix _summarized.xlsx.
"""
import argparse
import os
import time
from typing import Optional, Tuple, Dict

try:
    from dotenv import load_dotenv
    load_dotenv()  # Load .env file if it exists
except ImportError:
    pass  # dotenv not installed, skip

try:
    import openai
except Exception:
    openai = None

from openpyxl import load_workbook


def categorize_comment(text: str, categories: list, model: str = "gpt-4o") -> str:
    """Categorize the comment into one or more predefined categories using OpenAI.
    
    Returns a comma-separated string of categories.
    """
    if not text or not categories:
        return "Uncategorized"
    
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key or openai is None:
        return "Uncategorized"
    
    category_list = ", ".join(categories)
    prompt = f"""Analyze the following customer comment and assign it to one or more of the following categories.

Categories: [{category_list}]

- If the comment clearly fits one category, respond with that category.
- If it fits multiple, respond with a comma-separated list (e.g., "Product, Delivery").
- If it doesn't fit any, respond with "Other".

Comment: {text}

Category:"""
    
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=50,
            temperature=0.1,
        )
        category = resp.choices[0].message.content.strip()
        return category
    except Exception as e:
        print(f"Categorization failed: {e}")
        return "Uncategorized"


def analyze_overall_sentiment(
    positives_text: str,
    improvements_text: str,
    pos_score: float,
    imp_score: float,
    model: str = "gpt-4o"
) -> float:
    """Analyze overall sentiment using the categorized chunks and their scores.
    
    Takes the positive/improvement summaries and their individual sentiment scores,
    then asks LLM to determine an overall weighted sentiment.
    
    Returns a numeric sentiment score from 1-5.
    """
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key or openai is None:
        # Fallback: simple weighted average
        if pos_score == 0 and imp_score == 0:
            return 0
        elif pos_score == 0:
            return imp_score
        elif imp_score == 0:
            return pos_score
        else:
            return round((pos_score + imp_score) / 2, 1)
    
    prompt = f"""You are analyzing customer feedback that has been categorized into positive and negative aspects.

POSITIVE ASPECTS (Sentiment Score: {pos_score}/5):
{positives_text}

AREAS FOR IMPROVEMENT (Sentiment Score: {imp_score}/5):
{improvements_text}

Based on both the content and the individual sentiment scores, provide an OVERALL sentiment score that reflects:
1. The balance between positive and negative feedback
2. The intensity of each category (as shown by their scores)
3. The relative importance/emphasis in the original feedback

Respond with ONLY a single number from 1 to 5:
5 = Very Positive overall
4 = Positive overall
3 = Neutral/Mixed overall
2 = Negative overall
1 = Very Negative overall

Overall Score:"""
    
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=5,
            temperature=0.1,
        )
        score_str = resp.choices[0].message.content.strip()
        
        try:
            score = float(score_str)
            score = max(1, min(5, score))
            return score
        except ValueError:
            print(f"Could not parse overall score: {score_str}")
            # Fallback to weighted average
            if pos_score == 0 and imp_score == 0:
                return 0
            elif pos_score == 0:
                return imp_score
            elif imp_score == 0:
                return pos_score
            else:
                return round((pos_score + imp_score) / 2, 1)
            
    except Exception as e:
        print(f"Overall sentiment analysis failed: {e}")
        # Fallback to weighted average
        if pos_score == 0 and imp_score == 0:
            return 0
        elif pos_score == 0:
            return imp_score
        elif imp_score == 0:
            return pos_score
        else:
            return round((pos_score + imp_score) / 2, 1)


def analyze_sentiment(text: str, model: str = "gpt-4o") -> float:
    """Perform sentiment analysis on a text chunk using OpenAI.
    
    Returns a numeric sentiment score from 1-5:
    5 = Very Positive
    4 = Positive
    3 = Neutral
    2 = Negative
    1 = Very Negative
    """
    if not text or text.strip() in ["", "None mentioned", "No specific positives mentioned", "No specific improvements mentioned"]:
        return 0  # N/A cases get 0
    
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key or openai is None:
        return 0  # Fallback when API unavailable
    
    prompt = f"""Analyze the sentiment of the following text and provide a numeric sentiment score.

Respond with ONLY a single number from 1 to 5:
5 = Very Positive (highly enthusiastic, very satisfied)
4 = Positive (satisfied, good feedback)
3 = Neutral (mixed or neither positive nor negative)
2 = Negative (dissatisfied, criticism)
1 = Very Negative (very dissatisfied, serious complaints)

Text: {text}

Score:"""
    
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=5,
            temperature=0.1,
        )
        score_str = resp.choices[0].message.content.strip()
        
        # Parse the score
        try:
            score = float(score_str)
            # Clamp between 1-5
            score = max(1, min(5, score))
            return score
        except ValueError:
            print(f"Could not parse score: {score_str}")
            return 3  # Default to neutral if parsing fails
            
    except Exception as e:
        print(f"Sentiment analysis failed: {e}")
        return 0  # Return 0 on error


def analyze_sentiment_probs(text: str, model: str = "gpt-4o") -> Dict[str, float]:
    """Return probabilities for negative/neutral/positive that sum to 1.

    The result is a dict with keys: "negative", "neutral", "positive".
    Falls back to a simple heuristic if the API is unavailable or parsing fails.
    """
    # Default fallback: neutral when nothing to analyze
    if not text or text.strip() in ["", "None mentioned", "No specific positives mentioned", "No specific improvements mentioned"]:
        return {"negative": 0.0, "neutral": 1.0, "positive": 0.0}

    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key or openai is None:
        # Heuristic: mirror the numeric score from analyze_sentiment
        score = analyze_sentiment(text, model=model)
        if score == 0:
            return {"negative": 0.0, "neutral": 1.0, "positive": 0.0}
        if score <= 2:
            return {"negative": 0.8, "neutral": 0.2, "positive": 0.0}
        if score == 3:
            return {"negative": 0.2, "neutral": 0.6, "positive": 0.2}
        # 4 or 5
        return {"negative": 0.0, "neutral": 0.2, "positive": 0.8}

    prompt = f"""You are a sentiment classifier.

Given the following text, estimate the probabilities that the overall sentiment is
NEGATIVE, NEUTRAL, or POSITIVE.

Return a single line in the exact format:
negative=<number>, neutral=<number>, positive=<number>

Each number must be between 0 and 1 and they must sum to 1 (within rounding).

Text: {text}

Probabilities:"""

    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)

        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=30,
            temperature=0.1,
        )
        content = resp.choices[0].message.content.strip()

        # Expect: negative=0.1, neutral=0.2, positive=0.7
        probs = {"negative": 0.0, "neutral": 0.0, "positive": 0.0}
        try:
            parts = content.split(",")
            for part in parts:
                if "=" not in part:
                    continue
                k, v = part.split("=", 1)
                key = k.strip().lower()
                try:
                    val = float(v.strip())
                except ValueError:
                    continue
                if key in probs:
                    probs[key] = max(0.0, min(1.0, val))

            total = probs["negative"] + probs["neutral"] + probs["positive"]
            if total <= 0:
                raise ValueError("Sum of probabilities is zero")
            # Normalize so they sum to 1 exactly
            probs = {k: v / total for k, v in probs.items()}
            return probs
        except Exception as parse_err:
            print(f"Could not parse sentiment probabilities: '{content}' ({parse_err})")
    except Exception as e:
        print(f"Sentiment probabilities analysis failed: {e}")

    # Final fallback: neutral
    return {"negative": 0.0, "neutral": 1.0, "positive": 0.0}


def summarize_local(text: str) -> Tuple[str, str]:
    """Very small fallback summarizer: extracts positives and areas for improvement.
    
    Returns (positives_summary, improvements_summary)
    """
    if not text:
        return "", ""
    
    s = str(text).strip().lower()
    
    # Enhanced keyword-based heuristic with stronger positive signals
    strong_positive_indicators = ["great", "excellent", "love", "amazing", "wonderful", 
                                  "delighted", "fantastic", "outstanding", "passionate",
                                  "pleased", "satisfied", "happy", "thank you", "thanks"]
    positive_indicators = ["good", "helpful", "strength", "positive", "well", "best", 
                          "appreciate", "clear", "transparent", "flexible", "adaptable",
                          "smooth", "reliable", "quality", "professional"]
    negative_indicators = ["bad", "poor", "improve", "should", "could be better",
                          "issue", "problem", "concern", "weak", "lacking", "difficult",
                          "late", "delay", "not", "disappointing", "unfortunately"]
    
    # If the overall text is overwhelmingly positive, treat it as such
    strong_positive_count = sum(1 for word in strong_positive_indicators if word in s)
    negative_count = sum(1 for word in negative_indicators if word in s)
    
    # If highly positive with no negatives, put everything in positives
    if strong_positive_count >= 2 and negative_count == 0:
        summary = text.strip()
        if len(summary) > 100:
            summary = summary[:97] + "..."
        return summary, "None mentioned"
    
    sentences = []
    for sep in [". ", "! ", "? ", "\n"]:
        if sep in text:
            sentences = [s.strip() for s in text.split(sep) if s.strip()]
            break
    
    if not sentences:
        sentences = [text]
    
    positives = []
    improvements = []
    
    for sentence in sentences:
        sent_lower = sentence.lower()
        
        # Check for strong positive signals first
        has_strong_positive = any(word in sent_lower for word in strong_positive_indicators)
        has_positive = any(word in sent_lower for word in positive_indicators)
        has_negative = any(word in sent_lower for word in negative_indicators)
        
        # Prioritize strong positives
        if has_strong_positive and not has_negative:
            positives.append(sentence)
        elif has_positive and not has_negative:
            positives.append(sentence)
        elif has_negative and not has_strong_positive:
            improvements.append(sentence)
        elif not positives and not improvements:
            # Default to positives if unclear and no clear negatives
            positives.append(sentence)
    
    pos_summary = ". ".join(positives[:3]) if positives else "None mentioned"
    imp_summary = ". ".join(improvements[:2]) if improvements else "None mentioned"
    
    # Truncate if too long
    if len(pos_summary) > 150:
        pos_summary = pos_summary[:147] + "..."
    if len(imp_summary) > 150:
        imp_summary = imp_summary[:147] + "..."
    
    return pos_summary, imp_summary


def summarize_with_openai(text: str, model: str = "gpt-4o") -> Optional[Tuple[str, str]]:
    """Call OpenAI ChatCompletion to extract positives and areas for improvement.
    
    Returns (positives_summary, improvements_summary) or None if failed.
    Requires openai package and OPENAI_API_KEY env var set.
    """
    if openai is None:
        return None
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return None
    
    prompt = """Analyze the following comment carefully and categorize the feedback into four different categories and two distinct areas:

The four categories are:
1. Product
2. Service
3. Delivery
4. Payment

The two areas to focus on are:
1. POSITIVES: Extract ONLY genuinely positive feedback - things the person praised, appreciated, or said were done well. If the comment is purely critical or neutral, write "None mentioned".
2. IMPROVEMENTS: Extract ONLY constructive criticism, issues raised, or areas needing improvement. If the comment is purely positive with no concerns, write "None mentioned".

CRITICAL RULES:
- Do NOT convert positive statements into improvement areas
- Do NOT assume there must be problems if only praise is given
- Do NOT assume there must be positives if only criticism is given
- If a comment is entirely positive (e.g., "Great service, very satisfied"), put it in POSITIVES and put "None mentioned" for IMPROVEMENTS
- If a comment is entirely about problems (e.g., "Delivery was late, poor quality"), put "None mentioned" for POSITIVES

Format your response EXACTLY as:
POSITIVES: [your summary or "None mentioned"]
IMPROVEMENTS: [your summary or "None mentioned"]

Comment:
""" + str(text)
    
    try:
        # Use modern OpenAI API (v1.0+)
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=200,
            temperature=0.2,
        )
        content = resp.choices[0].message.content.strip()
        
        # Parse response more robustly
        positives = ""
        improvements = ""
        
        # Find the start of POSITIVES and IMPROVEMENTS
        pos_index = content.find("POSITIVES:")
        imp_index = content.find("IMPROVEMENTS:")

        # Determine the content of each section based on the indices
        if pos_index != -1 and imp_index != -1:
            if pos_index < imp_index:
                positives = content[pos_index + len("POSITIVES:"):imp_index].strip()
                improvements = content[imp_index + len("IMPROVEMENTS:"):].strip()
            else: # imp_index < pos_index
                improvements = content[imp_index + len("IMPROVEMENTS:"):pos_index].strip()
                positives = content[pos_index + len("POSITIVES:"):].strip()
        elif pos_index != -1:
            positives = content[pos_index + len("POSITIVES:"):].strip()
        elif imp_index != -1:
            improvements = content[imp_index + len("IMPROVEMENTS:"):].strip()

        # If parsing fails to find either, it's a malformed response.
        if not positives and not improvements:
             print("Warning: Malformed response from OpenAI, falling back to local summarizer.")
             return None # This will trigger the fallback in the run() function.
        
        return positives or "None mentioned", improvements or "None mentioned"
    except Exception as e:
        print(f"OpenAI call failed: {e}")
        return None


def run(
    path: str,
    sheet_name: Optional[str],
    input_col: str,
    positives_col: str,
    improvements_col: str,
    positives_sentiment_col: str,
    improvements_sentiment_col: str,
    overall_sentiment_col: str,
    raw_sentiment_col: str,
    category_col: str,
    categories: list,
    start_row: int,
    model: str,
    overwrite: bool,
    delay: float,
    force: bool = False,
):
    # Show which mode we're using
    api_key = os.getenv("OPENAI_API_KEY")
    if api_key and openai is not None:
        print(f"🤖 Using OpenAI API (model: {model})")
        print(f"   API Key: {api_key[:10]}...{api_key[-4:]}")
    else:
        print("⚠️  Using LOCAL summarizer (OpenAI not available)")
        if not api_key:
            print("   Reason: OPENAI_API_KEY not found in environment")
        if openai is None:
            print("   Reason: openai package not installed")
    print()
    
    wb = load_workbook(path)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    max_row = ws.max_row
    changed = 0
    for r in range(start_row, max_row + 1):
        in_cell = f"{input_col}{r}"
        pos_cell = f"{positives_col}{r}"
        imp_cell = f"{improvements_col}{r}"
        pos_sent_cell = f"{positives_sentiment_col}{r}"
        imp_sent_cell = f"{improvements_sentiment_col}{r}"
        overall_sent_cell = f"{overall_sentiment_col}{r}"
        raw_sent_cell = f"{raw_sentiment_col}{r}"
        category_cell = f"{category_col}{r}"
        
        comment = ws[in_cell].value
        if comment is None or str(comment).strip() == "":
            continue
        
        # Skip if already has all output cells populated (preserve existing)
        if not force and ws[pos_cell].value not in (None, "") and ws[imp_cell].value not in (None, "") and ws[category_cell].value not in (None, ""):
            continue

        positives = ""
        improvements = ""
        
        # Step 1: LLM splits into positive and negative parts
        if os.getenv("OPENAI_API_KEY") and openai is not None:
            result = summarize_with_openai(comment, model=model)
            if result:
                positives, improvements = result
            else:
                # Fallback to local if OpenAI fails
                positives, improvements = summarize_local(comment)
        else:
            positives, improvements = summarize_local(comment)

        # Step 2: Perform sentiment analysis on each chunk
        pos_sentiment = analyze_sentiment(positives, model=model)
        imp_sentiment = analyze_sentiment(improvements, model=model)
        
        # Step 3: Perform overall sentiment analysis using chunks + their scores
        overall_sentiment = analyze_overall_sentiment(
            positives, improvements, pos_sentiment, imp_sentiment, model=model
        )
        
        # Step 4: Perform raw sentiment analysis on original comment (no context)
        raw_sentiment = analyze_sentiment(comment, model=model)

        # Probability-style sentiment breakdowns (console only)
        pos_probs = analyze_sentiment_probs(positives, model=model)
        imp_probs = analyze_sentiment_probs(improvements, model=model)
        overall_probs = analyze_sentiment_probs(
            f"Positives: {positives}\nImprovements: {improvements}", model=model
        )
        raw_probs = analyze_sentiment_probs(comment, model=model)

        # Step 5: Categorize the comment
        category = categorize_comment(comment, categories, model=model)

        # Step 6: Write results to Excel
        ws[pos_cell].value = positives
        ws[imp_cell].value = improvements
        ws[pos_sent_cell].value = pos_sentiment
        ws[imp_sent_cell].value = imp_sentiment
        ws[overall_sent_cell].value = overall_sentiment
        ws[raw_sent_cell].value = raw_sentiment
        ws[category_cell].value = category
        
        changed += 1

        # Helper to map numeric score to label
        def _label(score: float) -> str:
            if score == 0:
                return "N/A"
            if score <= 2:
                return "Negative"
            if score == 3:
                return "Neutral"
            return "Positive"  # 4 or 5

        pos_label = _label(pos_sentiment)
        imp_label = _label(imp_sentiment)
        overall_label = _label(overall_sentiment)
        raw_label = _label(raw_sentiment)

        # Slide-friendly logging for each processed row
        print("\n" + "=" * 80)
        print(f"Row {r}")
        print("-" * 80)
        print(f"Original comment:\n{comment}\n")
        print(f"Detected category: {category}")
        print(f"Positives summary: {positives}")
        print(f"Improvements summary: {improvements}")
        print(
            f"Sentiment scores -> "
            f"Pos: {pos_sentiment} ({pos_label}) | "
            f"Imp: {imp_sentiment} ({imp_label}) | "
            f"Overall: {overall_sentiment} ({overall_label}) | "
            f"Raw: {raw_sentiment} ({raw_label})"
        )
        print("Probabilities (neg, neu, pos) ->")
        print(
            f"  Pos:  {pos_probs['negative']:.2f}, {pos_probs['neutral']:.2f}, {pos_probs['positive']:.2f}"
        )
        print(
            f"  Imp:  {imp_probs['negative']:.2f}, {imp_probs['neutral']:.2f}, {imp_probs['positive']:.2f}"
        )
        print(
            f"  Overall: {overall_probs['negative']:.2f}, {overall_probs['neutral']:.2f}, {overall_probs['positive']:.2f}"
        )
        print(
            f"  Raw:  {raw_probs['negative']:.2f}, {raw_probs['neutral']:.2f}, {raw_probs['positive']:.2f}"
        )
        print("=" * 80)
        
        if delay and os.getenv("OPENAI_API_KEY"):
            time.sleep(delay)

    if changed == 0:
        print("No comments found or no changes needed.")

    # Write output file
    dirname, basename = os.path.split(path)
    name, ext = os.path.splitext(basename)
    if overwrite:
        out_path = path
    else:
        out_path = os.path.join(dirname, f"{name}_summarized{ext}")
    
    # Try to save with helpful error message if file is locked
    try:
        wb.save(out_path)
        print(f"Saved {out_path} ({changed} rows updated)")
    except PermissionError:
        print(f"\n❌ ERROR: Cannot save to '{out_path}'")
        print("   The file is currently open in another program (Excel, etc.)")
        print("\n   Solutions:")
        print("   1. Close the Excel file and run the script again")
        print("   2. Run without --overwrite to create a new file instead:")
        print(f"      python scripts\\summarize_excel.py \"{path}\" --start-row {start_row}")
        raise


def parse_args():
    p = argparse.ArgumentParser(description="Summarize comments into positives and areas for improvement")
    p.add_argument("file", help="Path to the Excel file (.xlsx)")
    p.add_argument("--sheet", help="Sheet name (default: active)")
    p.add_argument("--input-col", default="B", help="Column with comments (default B)")
    p.add_argument("--positives-col", default="C", help="Column for positives summary (default C)")
    p.add_argument("--improvements-col", default="D", help="Column for improvements summary (default D)")
    p.add_argument("--positives-sentiment-col", default="E", help="Column for positives sentiment (default E)")
    p.add_argument("--improvements-sentiment-col", default="F", help="Column for improvements sentiment (default F)")
    p.add_argument("--overall-sentiment-col", default="G", help="Column for overall sentiment (default G)")
    p.add_argument("--raw-sentiment-col", default="H", help="Column for raw sentiment from original comment (default H)")
    p.add_argument("--category-col", default="I", help="Column for comment category (default I)")
    p.add_argument("--categories", default="Product,Service,Delivery,Payment", help="Comma-separated list of categories for classification")
    p.add_argument("--start-row", type=int, default=1, help="Row to start reading from (default 1)")
    p.add_argument("--model", default="gpt-4o", help="OpenAI model to use when available")
    p.add_argument("--overwrite", action="store_true", help="Overwrite the input file instead of creating _summarized")
    p.add_argument("--force", action="store_true", help="Force reprocessing of all rows, even if already summarized")
    p.add_argument("--delay", type=float, default=1.0, help="Seconds delay between OpenAI calls (default 1.0)")
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    try:
        run(
            path=args.file,
            sheet_name=args.sheet,
            input_col=args.input_col.upper(),
            positives_col=args.positives_col.upper(),
            improvements_col=args.improvements_col.upper(),
            positives_sentiment_col=args.positives_sentiment_col.upper(),
            improvements_sentiment_col=args.improvements_sentiment_col.upper(),
            overall_sentiment_col=args.overall_sentiment_col.upper(),
            raw_sentiment_col=args.raw_sentiment_col.upper(),
            category_col=args.category_col.upper(),
            categories=args.categories.split(','),
            start_row=args.start_row,
            model=args.model,
            overwrite=args.overwrite,
            delay=args.delay,
            force=args.force,
        )
    except Exception as e:
        print(f"Error: {e}")
        raise
